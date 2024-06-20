
#IMPORTS{

import vk_api
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType
import requests
from bs4 import BeautifulSoup as BS
import datetime
from openpyxl import *
import random
from vk_api import VkApi
from vk_api.utils import get_random_id

#}


# НАДО{

# vk_session = vk_api.VkApi(token='0eba16342f679f7fad7a17fdb0758a2144651be31cf45449dfc45f46f39f6d43cd4e2fc257bb275d7e0c5')


vkBotSession = VkApi(token='Тут ваш токен')
vk = vkBotSession.get_api()

longpoll = VkBotLongPoll(vkBotSession, 199029300 )

r = requests.get('https://sinoptik.ua/погода-ульяновск')
html = BS(r.content, 'html.parser')

wb = Workbook()
ws = wb.active

r = requests.get('https://sinoptik.ua/погода-ульяновск')
html = BS(r.content, 'html.parser')






wb_form = load_workbook(filename = 'Документ.xlsx')
wb_val = load_workbook(filename= 'Документ.xlsx', data_only = True)

sheet_form = wb_form['Sheet']
sheet_val = wb_val['Sheet']

Corona = 'https://www.worldometers.info/coronavirus/'
headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'}

full_page = requests.get(Corona,headers=headers)
soup = BS(full_page.content,'html.parser')

convert = soup.findAll("div",{"class":"maincounter-number"})






# online = vk_api.users.search(online=1)
# url = f"https://api.vk.com/method/users.search={online}&count=1&access_token={token}&v=5.52"
#}


#DEF{

def sender(id, text):
    vkBotSession.method('messages.send',{'chat_id' : id, 'message' : text, 'random_id' : 0})






#}


# OPTS}
opts = {
    "name": ('аркаша','эу','эй','аркадий','арк','арка'),

    "cmds": {

        "weather": ('погода'),

     }
}

#}

for event in longpoll.listen():
    if event.type == VkBotEventType.MESSAGE_NEW:
        if event.from_chat:
            if event.type == VkBotEventType.MESSAGE_NEW:  # проверка-и-отправка-смс
                peer = event.object.message['peer_id']  # id-чата
                ids = str(event.object.message['from_id'])  # id отправителя
            id = event.chat_id
            slovo = event.object.message['text'].lower()
            now = datetime.datetime.now()


            try:
                dey = event.message.action['type']
                invite_id = event.message.action['member_id']
            except:
                dey = ''
                invite_id = -100

                if dey == 'chat_invite_user':
                    sender(id,f'Приветствую тебя, @id{invite_id}!')


                s2 = ' ,:@.-()/?'

                for sim in s2:
                    slovo = slovo.replace(str(sim), ',')
                msg = slovo.split(',')
                while 1:
                    try:
                        msg.remove('')
                    except ValueError:
                        break
                print(msg)

                if msg[0] == 'аркаша':

                    if msg[1] == 'привет':
                        num = str(random.randint(2, 10))
                        a = 'A' + num
                        # print(a)

                        an = E5_val = sheet_val[a].value

                        sender(id, an)


                    elif msg[1] == 'как' and msg[2] == 'дела':

                        num = str(random.randint(2, 85))

                        a = 'B' + num

                        # print(a)

                        an = E5_val = sheet_val[a].value

                        sender(id, an)

                    elif msg[1] == 'что' and msg[2] == 'делаешь':
                        num = str(random.randint(2, 15))
                        a = 'C' + num
                        # print(a)

                        an = E5_val = sheet_val[a].value

                        sender(id, an)

                    elif msg[1] == 'корона':
                        sender(id,"Заражённых: " + convert[0].text)
                        sender(id,"Умерших: " + convert[1].text)
                        sender(id,"Выздоровевших: " + convert[2].text)




                    elif msg[1] == 'доллор':
                        # Основной класс
                        class Currency:
                            # Ссылка на нужную страницу
                            DOLLAR_RUB = 'https://www.google.com/search?sxsrf=ALeKk01NWm6viYijAo3HXYOEQUyDEDtFEw%3A1584716087546&source=hp&ei=N9l0XtDXHs716QTcuaXoAg&q=%D0%B4%D0%BE%D0%BB%D0%BB%D0%B0%D1%80+%D0%BA+%D1%80%D1%83%D0%B1%D0%BB%D1%8E&oq=%D0%B4%D0%BE%D0%BB%D0%BB%D0%B0%D1%80+&gs_l=psy-ab.3.0.35i39i70i258j0i131l4j0j0i131l4.3044.4178..5294...1.0..0.83.544.7......0....1..gws-wiz.......35i39.5QL6Ev1Kfk4'
                            # Заголовки для передачи вместе с URL
                            headers = {
                                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'}

                            current_converted_price = 0
                            difference = 5  # Разница после которой будет отправлено сообщение на почту

                            def __init__(self):
                                # Установка курса валюты при создании объекта
                                self.current_converted_price = float(self.get_currency_price().replace(",", "."))

                            # Метод для получения курса валюты
                            def get_currency_price(self):
                                # Парсим всю страницу
                                full_page = requests.get(self.DOLLAR_RUB, headers=self.headers)

                                # Разбираем через BeautifulSoup
                                soup = BS(full_page.content, 'html.parser')

                                # Получаем нужное для нас значение и возвращаем его
                                convert = soup.findAll("span",
                                                       {"class": "DFlfde", "class": "SwHCTb", "data-precision": 2})
                                return convert[0].text

                            # Проверка изменения валюты
                            def check_currency(self):
                                currency = float(self.get_currency_price().replace(",", "."))
                                if currency >= self.current_converted_price + self.difference:
                                    print("Курс сильно вырос, может пора что-то делать?")
                                    self.send_mail()
                                elif currency <= self.current_converted_price - self.difference:
                                    print("Курс сильно упал, может пора что-то делать?")
                                    self.send_mail()
                                sender(id,"Сейчас курс: 1 доллар = " + str(currency))


                        # Создание объекта и вызов метода
                        currency = Currency()
                        currency.check_currency()

                    elif msg[1] == 'евро':
                        # Основной класс
                        class Evro:
                            # Ссылка на нужную страницу
                            EVRO_RUB = 'https://www.google.com/search?sxsrf=ALeKk00bOlDaWVD6GynllhF3UknK8Jjk8g%3A1612117030165&ei=JvQWYJf6CMSrrgTS37CwAg&q=евро+к+рублю&oq=евк+рублю&gs_lcp=CgZwc3ktYWIQAxgAMgYIABAHEB4yBggAEAcQHjIGCAAQBxAeMgYIABAHEB4yBggAEAcQHjIGCAAQBxAeMgYIABAHEB4yBggAEAcQHjIGCAAQBxAeMgYIABAHEB46BwgjELADECc6BwgAELADEEM6CAgAEAcQChAeUI-JBFj_iwRgoJgEaARwAngAgAFLiAGSAZIBATKYAQCgAQGqAQdnd3Mtd2l6yAEKwAEB&sclient=psy-ab'
                            # Заголовки для передачи вместе с URL
                            headers = {
                                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'}

                            current_converted_price = 0
                            difference = 5  # Разница после которой будет отправлено сообщение на почту

                            def __init__(self):
                                # Установка курса валюты при создании объекта
                                self.current_converted_price = float(self.get_currency_price().replace(",", "."))

                            # Метод для получения курса валюты
                            def get_currency_price(self):
                                # Парсим всю страницу
                                full_page = requests.get(self.EVRO_RUB, headers=self.headers)

                                # Разбираем через BeautifulSoup
                                soup = BS(full_page.content, 'html.parser')

                                # Получаем нужное для нас значение и возвращаем его
                                convert = soup.findAll("span",
                                                       {"class": "DFlfde", "class": "SwHCTb", "data-precision": 2})
                                return convert[0].text

                            # Проверка изменения валюты
                            def check_currency(self):
                                currency = float(self.get_currency_price().replace(",", "."))
                                if currency >= self.current_converted_price + self.difference:
                                    print("Курс сильно вырос, может пора что-то делать?")
                                    self.send_mail()
                                elif currency <= self.current_converted_price - self.difference:
                                    print("Курс сильно упал, может пора что-то делать?")
                                    self.send_mail()
                                sender(id,"Сейчас курс: 1 евро = " + str(currency))


                        # Создание объекта и вызов метода
                        currency = Evro()
                        currency.check_currency()


                    elif msg[1] == 'погода':
                        for el in html.select('#content'):
                            t_min = el.select('.temperature .min')[0].text
                            t_max = el.select('.temperature .max')[0].text
                            text = el.select('.wDescription .description')[0].text

                        sender(id,"Привет, погода на сегодня:\n" +
                              t_min + ', ' + t_max + '\n' + text)



                elif msg[0] == 'его' and msg[1] == 'зовут' and msg[2] == 'аркаша':
                    sender(id,'Да, это я')

                elif msg[0] == 'погода':
                    for el in html.select('#content'):
                        t_min = el.select('.temperature .min')[0].text
                        t_max = el.select('.temperature .max')[0].text
                        text = el.select('.wDescription .description')[0].text

                    sender(id, "Привет, погода на сегодня:\n" +
                           t_min + ', ' + t_max + '\n' + text)

                elif msg[0] == 'время':
                    # os.system("start D:\Jarvis\Есть.wav")
                    sender(id,"Текушее время")
                    # print (now.strftime("%d-%m-%Y %H:%M"))
                    sender(id,now.strftime("%H:%M"))

                elif msg[0] == 'онлайн':
                    class VkOnline:
                        # Ссылка на нужную страницу
                        Online = 'https://www.google.com/search?sxsrf=ALeKk00bOlDaWVD6GynllhF3UknK8Jjk8g%3A1612117030165&ei=JvQWYJf6CMSrrgTS37CwAg&q=евро+к+рублю&oq=евк+рублю&gs_lcp=CgZwc3ktYWIQAxgAMgYIABAHEB4yBggAEAcQHjIGCAAQBxAeMgYIABAHEB4yBggAEAcQHjIGCAAQBxAeMgYIABAHEB4yBggAEAcQHjIGCAAQBxAeMgYIABAHEB46BwgjELADECc6BwgAELADEEM6CAgAEAcQChAeUI-JBFj_iwRgoJgEaARwAngAgAFLiAGSAZIBATKYAQCgAQGqAQdnd3Mtd2l6yAEKwAEB&sclient=psy-ab'
                        # Заголовки для передачи вместе с URL
                        headers = {
                            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'}

                        current_converted_price = 0
                        difference = 5  # Разница после которой будет отправлено сообщение на почту

                        def __init__(self):
                            # Установка курса валюты при создании объекта
                            self.current_converted_price = float(self.get_currency_price().replace(",", "."))

                        # Метод для получения курса валюты
                        def get_currency_price(self):
                            # Парсим всю страницу
                            full_page = requests.get(self.Online, headers=self.headers)

                            # Разбираем через BeautifulSoup
                            soup = BS(full_page.content, 'html.parser')

                            # Получаем нужное для нас значение и возвращаем его
                            convert = soup.findAll("span",
                                                   {"class": "page_top", "class": "_profile_online profile_online", "profile_online_lv": 2})
                            return convert[0].text

                        # Проверка изменения валюты
                        def check_currency(self):
                            currency = float(self.get_currency_price().replace(",", "."))
                            if currency >= self.current_converted_price + self.difference:
                                print("Курс сильно вырос, может пора что-то делать?")
                                self.send_mail()
                            elif currency <= self.current_converted_price - self.difference:
                                print("Курс сильно упал, может пора что-то делать?")
                                self.send_mail()
                            sender(id, "Сейчас курс: 1 евро = " + str(currency))


                    # Создание объекта и вызов метода
                    currency = VkOnline()
                    currency.check_currency()




                # elif msg[0] == 'онлайн':
                #     sender(id, online)
                    # elif msg[1] == 'как дела' or msg == 'как дела?':
                    #     sender(id,'Хорошо')




















