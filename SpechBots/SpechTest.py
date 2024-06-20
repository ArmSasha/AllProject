#!/usr/bin/env python
# -*- coding: utf-8 -*-
import pyttsx3
from openpyxl import *
import random
import os
import webbrowser as web
import datetime
import time
import winsound
from colorama import init

from colorama import Fore, Back, Style
import requests
from bs4 import BeautifulSoup as BS
import wikipedia

import speech_recognition as sr
import sys


init() # COLORAMA INIT

wikipedia.set_lang("ru")

engline = pyttsx3.init()
voices = engline.getProperty('voices')
engline.setProperty('voice', 'eu')

wb = Workbook()
ws = wb.active



wb_form = load_workbook(filename = 'Документ.xlsx')
wb_val = load_workbook(filename= 'Документ.xlsx', data_only = True)

sheet_form = wb_form['Sheet']
sheet_val = wb_val['Sheet']


keys = {
	'name': ['как тебя зовут', 'имя', 'зовут как', 'назови имя']

}









#_______________________________________________________

def word(x):
    engline.say(x)
    engline.runAndWait()


#_______________________________________________________

for voice in voices:
    if voice.name == 'IVONA 2 Tatyana OEM':
        engline.setProperty('voice', voice.id)





print("Привет, чем я могу помочь вам?")
word("Привет, чем я могу помочь вам?")

#_______________________________________________________

def command():
	# Создаем объект на основе библиотеки
	# speech_recognition и вызываем метод для определения данных
	r = sr.Recognizer()

	# Начинаем прослушивать микрофон и записываем данные в source
	with sr.Microphone() as source:
		# Просто вывод, чтобы мы знали когда говорить
		print("Говорите")
		# Устанавливаем паузу, чтобы прослушивание
		# началось лишь по прошествию 1 секунды
		r.pause_threshold = 1
		# используем adjust_for_ambient_noise для удаления
		# посторонних шумов из аудио дорожки
		r.adjust_for_ambient_noise(source, duration=1)
		# Полученные данные записываем в переменную audio
		# пока мы получили лишь mp3 звук
		audio = r.listen(source)

	try: # Обрабатываем все при помощи исключений
		""" 
		Распознаем данные из mp3 дорожки.
		Указываем что отслеживаемый язык русский.
		Благодаря lower() приводим все в нижний регистр.
		Теперь мы получили данные в формате строки,
		которые спокойно можем проверить в условиях
		"""
		zadanie = r.recognize_google(audio, language="ru-RU").lower()
		# Просто отображаем текст что сказал пользователь
		print("Вы сказали: " + zadanie)
	# Если не смогли распознать текст, то будет вызвана эта ошибка
	except sr.UnknownValueError:
		# Здесь просто проговариваем слова "Я вас не поняла"
		# и вызываем снова функцию command() для
		# получения текста от пользователя
		print("Я вас не поняла")
		word("Я вас не поняла")
		zadanie = command()

	# В конце функции возвращаем текст задания
	# или же повторный вызов функции
	return zadanie

#_______________________________________________________

def makeSomething(zadanie):
	# Попросту проверяем текст на соответствие
	# Если в тексте что сказал пользователь есть слова
	# "открыть сайт", то выполняем команду
	if 'открыть сайт' in zadanie:
		# Проговариваем текст
		print("Уже открываю")
		word("Уже открываю")
		# Указываем сайт для открытия
		url = zadanie
		# Открываем сайт
		webbrowser.open(url)

#_______________________________________________________

	elif 'стоп' in zadanie:
		# Проговариваем текст
		print("Да, конечно, без проблем")
		word("Да, конечно, без проблем")
		# Выходим из программы
		sys.exit()

#_______________________________________________________

	elif 'имя' in zadanie:
		print("Меня зовут Сири")
		word("Меня зовут Сири")

#_______________________________________________________

	elif 'привет' in zadanie:
		num = str(random.randint(2,10))
		a = 'A' + num

		an = E5_val = sheet_val[a].value

		print(colorama.Fore.LIGHTCYAN_EX + an)
		word(colorama.Fore.LIGHTCYAN_EX + an)

#_______________________________________________________

	elif 'как дела' in zadanie:
		num = str(random.randint(2, 85))
		a = 'B' + num

		an = E5_val = sheet_val[a].value

		print(colorama.Fore.LIGHTCYAN_EX + an)
		word(colorama.Fore.LIGHTCYAN_EX + an)

#_______________________________________________________

	elif 'что делаешь' in zadanie:
		num = str(random.randint(2, 15))
		a = 'C' + num

		an = E5_val = sheet_val[a].value

		print(colorama.Fore.LIGHTCYAN_EX + an)
		word(colorama.Fore.LIGHTCYAN_EX + an)

#_______________________________________________________

	elif 'сколько тебе лет' in zadanie:
		num = str(random.randint(2, 15))
		a = 'D' + num

		an = E5_val = sheet_val[a].value

		print(colorama.Fore.LIGHTCYAN_EX + an)
		word(colorama.Fore.LIGHTCYAN_EX + an)

#_______________________________________________________

	# elif 'аникдот' in zadanie:
	# 	url = "https://www.anekdot.ru/random"

	# 	request = requests.get(url)

	# 	soup = BeautifulSoup(request.text. "html.parser")

	# 	anekdot = soup.find_all("td", class_="title")

	# 	for temes in anekdot:

	# 		temes = temes.find("a", {'class':'storylink'})

	# 		if temes is not None:
	# 			print(temes.text)
	# 	# z=''
	# 	# s=requests.get('https://www.anekdot.ru/random/')
	# 	# b=bs4.BeautifulSoup(s.text, "html.parser")
	# 	# p=b.select('.a_rnd')
	# 	# for x in p:
	# 	#     s=(x.getText().strip())

	# 	print(s)
	# 	word(s)
	elif 'узнай' in zadanie:

		text_to_search = zadanie
		search_results = wikipedia.search(text_to_search, results=5)

		if len(search_results) == 0 or text_to_search == 'exit':
			print(f"По запросу '{text_to_search}' ничего не найдено")
			exit()

		for index, result in enumerate(search_results):
			print(f"{index}) {result}")

		get_one = input("Номер результата: ")
		search_result = search_results[int(get_one)]
		text = wikipedia.summary(search_result)

		print(text)
		word(text)

		# python_page = wikipedia.page(zadanie)

		# # print(colorama.Fore.LIGHTCYAN_EX + python_page.html)
		# # word(colorama.Fore.LIGHTCYAN_EX + python_page.html)
		# # print(python_page.html)
		# # word(python_page.html)

		# print(Fore.GREEN + python_page.original_title)
		# word(python_page.original_title)

		# print(Fore.YELLOW + python_page.summary)
		# word(python_page.summary)



# Вызов функции для проверки текста будет
# осуществляться постоянно, поэтому здесь
# прописан бесконечный цикл while
while True:
	makeSomething(command())