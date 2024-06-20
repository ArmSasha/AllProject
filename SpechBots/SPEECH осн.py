#!/usr/bin/env python
# -*- coding: utf-8 -*-
import bs4
import pyttsx3
from openpyxl import *
import random
import os
import webbrowser as web
import datetime
import time
import winsound
import colorama
import requests
from bs4 import BeautifulSoup as BS


engline = pyttsx3.init()
voices = engline.getProperty('voices')
engline.setProperty('voice', 'ru')

wb = Workbook()
ws = wb.active

r = requests.get('https://sinoptik.ua/погода-ульяновск')
html = BS(r.content, 'html.parser')


wb_form = load_workbook(filename = 'Документ.xlsx')
wb_val = load_workbook(filename= 'Документ.xlsx', data_only = True)

sheet_form = wb_form['Sheet']
sheet_val = wb_val['Sheet']

Corona = 'https://www.worldometers.info/coronavirus/'
headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'}

full_page = requests.get(Corona,headers=headers)
soup = BS(full_page.content,'html.parser')

convert = soup.findAll("div",{"class":"maincounter-number"})



def word(x):
    engline.say(x)
    engline.runAndWait()


program = True

for voice in voices:
    if voice.name == 'IVONA 2 Tatyana OEM':
        engline.setProperty('voice', voice.id)



while program:

    # for  i in range(1):
    #     word('Привет! Меня завут Пятница, как зовут тебя')

    slovo = input().lower()
    now = datetime.datetime.now()

    s2 = ' ,:@.-()/?'

    for sim in s2:
        slovo = slovo.replace(str(sim), ',')
    msg = slovo.split(',')
    while 1:
        try:
            msg.remove('')
        except ValueError:
            break
    print(msg)


    if msg[0] == 'привет':
        num = str(random.randint(2,10))
        a = 'A'+num
        # print(a)

        an = E5_val = sheet_val[a].value

        print(colorama.Fore.LIGHTCYAN_EX + an)

    # if k == 'привет':
    #     if now.hour > 12:
    #         word('И тебе привет')
    #     if now.hour < 12:
    #         # mus = pyglet.resource.media("Доброеутро.mp3")
    #         # mus.play()
    #         # pyglet.app.run()
    #         Sp = "D:\SPEECH\Jarvis\Доброеутро.wav"
    #         winsound.PlaySound(Sp, winsound.SND_FILENAME)
    #         #os.system("start D:\SPEECH\Jarvis\Доброеутро.wav")
    if msg[0] == 'пока':
        word('До скорых встреч')
        program = False

    elif msg[0] == 'как' and msg[1] == 'дела':
        num = str(random.randint(2, 85))
        a = 'B' + num

        an = E5_val = sheet_val[a].value

        print(colorama.Fore.LIGHTCYAN_EX + an)



    elif msg[0] == 'что' and msg[1] == 'делаешь':
        num = str(random.randint(2, 15))
        a = 'C' + num

        an = E5_val = sheet_val[a].value

        print(colorama.Fore.LIGHTCYAN_EX + an)

    elif msg[0] == 'сколько' and msg[1] == 'тебе' and msg[2] == 'лет':
        word('Я ещё молодая')

    elif msg[0] == 'что' and msg[1] == 'ты' and msg[2] =='умеешь':
        word('моногое, по пальцам не сосчитать')

    elif msg[0] == 'как' and msg[1] == 'тебя' and msg[2] == 'зовут':
        word('Ирина, но у меня есть друг Джарвис, а тебя?')




#           АНИКДОТ{
    elif msg[0] == 'раскажи' and msg[1] == 'аникдот' or msg[0] == 'аникдот' :
        z=''
        s=requests.get('https://www.anekdot.ru/random/anekdot/')
        b=bs4.BeautifulSoup(s.text, "html.parser")
        p=b.select('.text')
        for x in p:
            s=(x.getText().strip())

        print(s)
        word(s)

#}

    #if k == 'аник':
        #Dobr = "D:\SPEECH\Jarvis\Доброеутро.wav"
        #winsound.PlaySound(Dobr, winsound.SND_FILENAME)




#           ЗАПРОСЫ{

    elif msg[0] == 'корона':
        print("Заражённых: " + convert[0].text)
        print("Умерших: " + convert[1].text)
        print("Вылечиных: " + convert[2].text)

    if msg[0] == 'курс':
        print('Выберите\n'
              'доллор \n'
              'евро')
    elif msg[0] == 'доллор':
        # Основной класс
        class Currency:
            # Ссылка на нужную страницу
            DOLLAR_RUB = 'https://www.google.com/search?sxsrf=ALeKk01NWm6viYijAo3HXYOEQUyDEDtFEw%3A1584716087546&source=hp&ei=N9l0XtDXHs716QTcuaXoAg&q=%D0%B4%D0%BE%D0%BB%D0%BB%D0%B0%D1%80+%D0%BA+%D1%80%D1%83%D0%B1%D0%BB%D1%8E&oq=%D0%B4%D0%BE%D0%BB%D0%BB%D0%B0%D1%80+&gs_l=psy-ab.3.0.35i39i70i258j0i131l4j0j0i131l4.3044.4178..5294...1.0..0.83.544.7......0....1..gws-wiz.......35i39.5QL6Ev1Kfk4'
            # Заголовки для передачи вместе с URL
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'}

            current_converted_price = 0
            difference = 5  # Разница после которой будет отправлено сообщение на почту

            def __init__(self):
                # Установка курса валюты при создании объекта
                self.current_converted_price = float(self.get_currency_price().replace(",", "."))

            # Метод для получения курса валюты
            def get_currency_price(self):
                # Парсим всю страницу
                full_page = requests.get(self.DOLLAR_RUB, headers=self.headers)

                # Разбираем через BeautifulSoup
                soup = BS(full_page.content, 'html.parser')

                # Получаем нужное для нас значение и возвращаем его
                convert = soup.findAll("span", {"class": "DFlfde", "class": "SwHCTb", "data-precision": 2})
                return convert[0].text

            # Проверка изменения валюты
            def check_currency(self):
                currency = float(self.get_currency_price().replace(",", "."))
                if currency >= self.current_converted_price + self.difference:
                    print("Курс сильно вырос, может пора что-то делать?")
                    self.send_mail()
                elif currency <= self.current_converted_price - self.difference:
                    print("Курс сильно упал, может пора что-то делать?")
                    self.send_mail()
                print("Сейчас курс: 1 доллар = " + str(currency))

        # Создание объекта и вызов метода
        currency = Currency()
        currency.check_currency()

    elif msg[0] == 'евро':
        # Основной класс
        class Evro:
            # Ссылка на нужную страницу
            EVRO_RUB = 'https://www.google.com/search?sxsrf=ALeKk00bOlDaWVD6GynllhF3UknK8Jjk8g%3A1612117030165&ei=JvQWYJf6CMSrrgTS37CwAg&q=евро+к+рублю&oq=евк+рублю&gs_lcp=CgZwc3ktYWIQAxgAMgYIABAHEB4yBggAEAcQHjIGCAAQBxAeMgYIABAHEB4yBggAEAcQHjIGCAAQBxAeMgYIABAHEB4yBggAEAcQHjIGCAAQBxAeMgYIABAHEB46BwgjELADECc6BwgAELADEEM6CAgAEAcQChAeUI-JBFj_iwRgoJgEaARwAngAgAFLiAGSAZIBATKYAQCgAQGqAQdnd3Mtd2l6yAEKwAEB&sclient=psy-ab'
            # Заголовки для передачи вместе с URL
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'}

            current_converted_price = 0
            difference = 5  # Разница после которой будет отправлено сообщение на почту

            def __init__(self):
                # Установка курса валюты при создании объекта
                self.current_converted_price = float(self.get_currency_price().replace(",", "."))

            # Метод для получения курса валюты
            def get_currency_price(self):
                # Парсим всю страницу
                full_page = requests.get(self.EVRO_RUB, headers=self.headers)

                # Разбираем через BeautifulSoup
                soup = BS(full_page.content, 'html.parser')

                # Получаем нужное для нас значение и возвращаем его
                convert = soup.findAll("span", {"class": "DFlfde", "class": "SwHCTb", "data-precision": 2})
                return convert[0].text

            # Проверка изменения валюты
            def check_currency(self):
                currency = float(self.get_currency_price().replace(",", "."))
                if currency >= self.current_converted_price + self.difference:
                    print("Курс сильно вырос, может пора что-то делать?")
                    self.send_mail()
                elif currency <= self.current_converted_price - self.difference:
                    print("Курс сильно упал, может пора что-то делать?")
                    self.send_mail()
                print("Сейчас курс: 1 евро = " + str(currency))

        # Создание объекта и вызов метода
        currency = Evro()
        currency.check_currency()

    elif msg[0] == 'погода':
        for el in html.select('#content'):
            t_min = el.select('.temperature .min')[0].text
            t_max = el.select('.temperature .max')[0].text
            text = el.select('.wDescription .description')[0].text

        word("Привет, погода на сегодня:\n" +
               t_min + ', ' + t_max + '\n' + text)

    elif msg[0] == 'google':
        os.system("start C://Users/Саша/Desktop/Google.lnk")

    elif msg[0] == 'youtube':
        web.open("https://www.youtube.com/")

    elif msg[0] == 'учи ру':
        web.open("https://uchi.ru/students/main")

    elif msg[0] == 'открой' or msg[0] == 'найди' or msg[0] == 'запроси' or msg[0] == 'запрос' or msg[0] == 'открыть':
        word('Введите запрос')
        print("Введите запрос:")
        zapr = input()

        if zapr != 'отмена':
            Sp = "D:\SPEECH\Jarvis\Загружаюсэр.wav"
            winsound.PlaySound(Sp, winsound.SND_FILENAME)

            web.open("https://yandex.ru/yandsearch?clid=2028026&text={}&tr=11373".format(zapr))

        else:
            print("Запрос отменён, введите другую команду:")
            Sp = "D:\SPEECH\Jarvis\Есть.wav"
            winsound.PlaySound(Sp, winsound.SND_FILENAME)

#{



#ВРЕМЯ{
    elif msg[0] == 'время':
        # os.system("start D:\Jarvis\Есть.wav")
        print("Текушее время")
        # print (now.strftime("%d-%m-%Y %H:%M"))
        print(now.strftime("%H:%M"))

        word(now.strftime("%H:%M"))
    # if k == 'сколько':
    #     mus = pyglet.resource.media("Доброеутро.mp3")
    #     mus.play()
    #     pyglet.app.run()
#}




#ПРОГРАММЫ {
    elif msg[0] == 'взлом':
        vrem = int(input("Введите время работы взлома:"))
        os.system("start D:\SPEECH\Matrisa.bat")
        time.sleep(vrem)
        os.system("TASKKILL /F /IM cmd.exe")
        # sys.exit()
    '''
    
     0 = Черный      8 = Серый
     1 = Синий       9 = Светло-синий
     2 = Зеленый     A = Светло-зеленый
     3 = Голубой     B = Светло-голубой
     4 = Красный     C = Светло-красный
     5 = Лиловый     D = Светло-лиловый
     6 = Желтый      E = Светло-желтый
     7 = Белый       F = Ярко-белый
    '''

    # if k == 'starwars':
    #     os.system('start D:\SPEECH\SPEECH\StarWars.bat')

    if msg[0] == 'часы':
        os.system("start D:\SPEECH\F\dist\F\F.exe")

    elif msg[0] == 'таймер':
        os.system("start D:\SPEECH\Timer\dist\Timer\Timer.exe")

    elif msg[0] == 'секундомер':
        os.system("start D:\SPEECH\Sekond\dist\Sekond\Sekond.exe")

    elif msg[0] == 'матрица':
        os.system("start D:\SPEECH\Matric\dist\Matric\Matric.exe")

    elif msg[0] == 'хак':
        os.system("start D:\SPEECH\Взлом.bat")

    elif msg[0] == 'тетрис':
        os.system("start D:\SPEECH\Tetris\dist\Tetris\Tetris.exe")

    elif msg[0] == 'лабиринт':
        os.system("start D:\SPEECH\Maze\dist\Maze\Maze.exe")

    elif msg[0] == 'аркада':
        os.system("start D:\SPEECH\Arkad\dist\Arkad\Arkad.exe")

    elif msg[0] == 'aeroblaster':
        os.system("start D:\SPEECH\Aeroblaster\Aeroblaster.exe")

    elif msg[0] == 'cranberryWoodlands':
        os.system("start D:\SPEECH\CranberryWoodlands\CranberryWoodlands.exe")

    elif msg[0] == 'frog':
        os.system("start D:\SPEECH\Frog\FroggosAdventure.exe")

    elif msg[0] == 'lifelust':
        os.system("start D:\SPEECH\Lifelust\Lifelust.exe")

    elif msg[0] == 't-rex':
        os.system('start D:/SPEECH/T-rex/"T-rexgame+.exe"')

    elif msg[0] == 'играна1или2':
        os.system('start D:/SPEECH/Играна1или2/"Dungeons And Caves.exe"')

    elif msg[0] == 'на2':
        os.system("start D:\SPEECH\На2\Catchio.exe")

    elif msg[0] == 'не понятно что':
        os.system("start D:/SPEECH/Непонятночто/OPEN-GAME.exe")

    elif msg[0] == 'персы':
        os.system("start D:/SPEECH/Персы/dressup.exe")

    elif msg[0] == 'kliker':
        os.system("start D:\SPEECH\Kliker\dist\Kliker\Kliker.exe")
# }




# ИГРЫ {
    if msg[0] == 'игры':
        print("1)тетрис \
2)лабиринт \
3)аркада \
4)aeroblaster \
5)cranberryWoodlands \
6)frog \
7)lifelust \
8)t-rex \
9)играна1или2 \
10)на2 \
11)не понятно что \
12)персы \
13)kliker")
        play = input().strip().lower()
        print(play)
        if play == 'отмена':
            print("Отменено введите другую команду")
            k = input().lower().strip()
        if play == str(1):
            os.system("start D:\SPEECH\Tetris\dist\Tetris\Tetris.exe")

        if play == str(2):
            os.system("start D:\SPEECH\Maze\dist\Maze\Maze.exe")

        if play == str(3):
            os.system("start D:\SPEECH\Arkad\dist\Arkad\Arkad.exe")

        if play == str(4):
            os.system("start D:\SPEECH\Aeroblaster\Aeroblaster.exe")

        if play == str(5):
            os.system("start D:\SPEECH\CranberryWoodlands\CranberryWoodlands.exe")

        if play == str(6):
            os.system("start D:\SPEECH\Frog\FroggosAdventure.exe")

        if play == str(7):
            os.system("start D:\SPEECH\Lifelust\Lifelust.exe")

        if play == str(8):
            os.system('start D:/SPEECH/T-rex/"T-rexgame+.exe"')

        if play == str(9):
            os.system('start D:/SPEECH/Играна1или2/"Dungeons And Caves.exe"')

        if play == str(10):
            os.system("start D:\SPEECH\На2\Catchio.exe")

        if play == str(11):
            os.system("start D:/SPEECH/Непонятночто/OPEN-GAME.exe")

        if play == str(12):
            os.system("start D:/SPEECH/Персы/dressup.exe")

        if play == str(13):
            os.system("start D:\SPEECH\Kliker\dist\Kliker\Kliker.exe")

# }
















